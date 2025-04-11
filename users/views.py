import csv
import json

import openpyxl
from django.contrib import messages
from django.contrib.auth import authenticate, logout, login
from django.contrib.auth.models import User
from django.contrib.sites.shortcuts import get_current_site
from django.core.mail import EmailMessage
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils.encoding import force_bytes
from django.utils.encoding import force_str
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.views import View
from django.views.generic import FormView
from django.views.generic import ListView, CreateView, UpdateView, DeleteView

from app.forms import CustomerForm
from users.forms import RegisterModelForm, LoginForm
from users.models import Customer
from users.tokens import account_activation_token


# Create your views here.


class CustomerListView(ListView):
    model = Customer
    template_name = 'app/customers.html'
    context_object_name = 'customers'

    def get_queryset(self):
        query = self.request.GET.get('q', '')
        if query:
            return Customer.objects.filter(name__icontains=query)
        return Customer.objects.all()


class CustomerDetailView(View):
    def get(self, request, slug):
        customer = get_object_or_404(Customer, slug=slug)
        return render(request, 'app/customer-details.html', {'customer': customer})


class CustomerCreateView(CreateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'app/customer_add.html'
    success_url = reverse_lazy('users:customers')


class CustomerUpdateView(UpdateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'app/customer_update.html'
    success_url = reverse_lazy('users:customers')

    def get_object(self, queryset=None):
        return get_object_or_404(Customer, name=self.kwargs.get('slug'))


class CustomerDeleteView(DeleteView):
    model = Customer
    template_name = 'app/customer_delete.html'
    success_url = reverse_lazy('users:customers')

    def get_object(self, queryset=None):
        return get_object_or_404(Customer, slug=self.kwargs.get('slug'))


def login_page(request):
    form = LoginForm()
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            user = authenticate(request, email=cd['email'], password=cd['password'])
            if user:
                if user.is_active:
                    login(request, user)
                    return redirect('app:index')
                else:
                    messages.add_message(
                        request,
                        messages.ERROR,
                        'Disabled account'
                    )
                    return render(request, 'users/login.html')
            else:
                messages.add_message(
                    request,
                    messages.ERROR,
                    'Username or Password invalid'
                )
                return render(request, 'users/login.html')

    return render(request, 'users/login.html', {'form': form})


class RegisterPage(FormView):
    template_name = 'users/register.html'
    form_class = RegisterModelForm
    success_url = reverse_lazy('users:verify_email_done')

    def form_valid(self, form):
        user = form.save(commit=False)
        user.is_staff = True
        user.is_superuser = True
        user.is_active = False
        user.set_password(user.password)
        user.save()
        current_site = get_current_site(self.request)
        subject = "Verify Email"
        message = render_to_string('users/Email/verify_email_message.html', {
            'request': self.request,
            'user': user,
            'domain': current_site.domain,
            'uid': urlsafe_base64_encode(force_bytes(user.pk)),
            'token': account_activation_token.make_token(user),
        })
        user.save()
        email = EmailMessage(
            subject, message, to=[user.email]
        )
        email.content_subtype = 'html'
        email.send()
        return super().form_valid(form)


def email_required(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        user = request.user

        if user.is_authenticated and not user.is_active:
            user.email = email
            user.save()
            return redirect('app:index')
    return render(request, 'Github/email-required.html')


def verify_email_done(request):
    return render(request, 'users/Email/verify_email_done.html')


def verify_email_confirm(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = Customer.objects.get(pk=uid)

    except(TypeError, ValueError, OverflowError, Customer.DoesNotExist):
        user = None
    if user and account_activation_token.check_token(user, token):
        user.is_active = True
        user.save()
        login(request, user, backend='django.contrib.auth.backends.ModelBackend')
        messages.success(request, 'Your email has been verified.')
        return redirect('app:index')
    else:
        messages.warning(request, 'The link is invalid.')
    return render(request, 'users/email/verify_email_confirm.html')



class LogoutView(View):
    def get(self, request):
        return render(request, 'users/logout.html')

    def post(self, request):
        logout(request)
        messages.success(request, "Succesfully logout")
        return redirect('app:index')


def export_data(request):
    format = request.GET.get('format')
    if format == 'csv':
        meta = Customer._meta
        field_names = [field.name for field in meta.fields]
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="customer_list.csv'
        writer = csv.writer(response)
        writer.writerow(field_names)
        for obj in Customer.objects.all():
            row = writer.writerow([getattr(obj, field) for field in field_names])
            return response

    elif format == 'json':
        response = HttpResponse(content_type='application/json')
        data = list(Customer.objects.all().values_list('id', 'name', 'email', 'phone', 'billing_address', 'password',
                                                       'created_at', 'image', 'slug', 'VAT_Number'))
        response.write(json.dumps(data, indent=4, default=str))
        response['Content-Disposition'] = 'attachment; filename="customer.json"'
        return response

    elif format == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Customers List'

        colums = ['id', 'name', 'email', 'phone', 'billing_address', 'password', 'created_at', 'image', 'slug',
                  'VAT_Number']
        fields = ['id', 'name', 'email', 'phone', 'billing_address', 'password', 'created_at', 'image', 'slug',
                  'VAT_Number']

        for col_num, column_title in enumerate(colums, 1):
            ws.cell(row=1, column=col_num, value=column_title)

        customers = Customer.objects.all().values_list(*fields)

        for row_num, row_data in enumerate(customers, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                if hasattr(cell_value, 'isoformat'):
                    cell_value = cell_value.isoformat()
                elif hasattr(cell_value, 'url'):
                    cell_value = cell_value.url
                ws.cell(row=row_num, column=col_num, value=cell_value)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="customer.xlsx"'
        wb.save(response)
        return response

    else:
        response = HttpResponse(status=404)
        response.content = 'Bad request'
        return response

def home(request):
    if request.device_type == "Mobile":
        # return render(request, "mobile/home.html")     Mobile Telefon uchun html yaratiladi. Tez kunda!!!
        pass
    else:
        return render(request, "app/index.html")



